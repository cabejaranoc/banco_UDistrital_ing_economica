import tkinter as tk
from tkinter import ttk
from tkinter import font as tkfont
from Finanzas.interest import * 
import openpyxl
from openpyxl.styles import Alignment
from tkinter import messagebox

class MainWindow:

    def __init__(self, root):
        self.root = root
        self.root.title("Calculadora de Ingeniería Económica")
        self.Mitad_pantalla(500,800,self.root)
        self.root.configure(bg='LightSkyBlue')
        self.A =0
        self.a = 0
        self.i = 0
        self.n = 0
        self.S=0
        self.banderaS = False
        self.banderaa = False        

        self.menu_principal()

    def menu_principal(self):
        Title_font = tkfont.Font(family="Verdana", size=16, weight="bold")
        Btn_font = tkfont.Font(family="Verdana",size=12)
        style = ttk.Style()
        style.configure('TButton', font=Btn_font)       

        label = ttk.Label(self.root, text="Seleccione una funcionalidad",font=Title_font,background='LightSkyBlue')
        label.pack(pady=70)

        ICButton = ttk.Button(self.root, width=25, style='TButton', text="Interés Compuesto", command=self.CompuestoMain)
        ICButton.pack(pady=10)

        EQTButton = ttk.Button(self.root, width=25, style='TButton', text="Equivalencia de Tasas",command=self.EQtasas)
        EQTButton.pack(pady=10)

        AButton = ttk.Button(self.root, width=25, style='TButton', text="Anualidades", command=self.AnualidadMain)
        AButton.pack(pady=10)

        TablesButton = ttk.Button(self.root, width=25, style='TButton', text="Generar Tablas", command=self.Tables)
        TablesButton.pack(pady=10)

        # Labels para los nombres de los autores
        authors = ["Autores","Robinson Stiven Inagan Ochoa - 20232678006", "Brahian Estiven Ruiz Murcia - 20241678037", "Yiver Stiven Moreno Parra - 20241678029", "Ana Raquel Molano Ramos - 20222375023"]
        
        for author in authors:
            author_label = ttk.Label(self.root, text=author, font=Btn_font, background='LightSkyBlue')
            author_label.pack(pady=5)          

    def CompuestoMain(self):
        self.root.withdraw()
        main_window = tk.Toplevel()
        main_window.title("Cálculo de Interés Compuesto")
        self.Mitad_pantalla(300,400,main_window)

        main_window.configure(bg='LightSkyBlue')
        # Combobox para selección de opción
        ttk.Label(main_window, text="Seleccione el cálculo:", background='LightSkyBlue').pack(pady=5)
        options = ["valor", "tiempo", "tasa"]
        option_combobox = ttk.Combobox(main_window, values=options, state="readonly")
        option_combobox.pack(pady=5)

        # Botón Continuar
        continue_button = ttk.Button(main_window, text="Continuar", 
                                    command=lambda: self.ValidarOpcionComp(option_combobox.get(), main_window))
        continue_button.pack(pady=10)

        # Botón Volver
        back_button = ttk.Button(main_window, text="Volver", command=lambda: self.Volver(main_window))
        back_button.pack(pady=10)

        main_window.protocol("WM_DELETE_WINDOW", lambda: self.Volver(main_window))

    def CompuestoN(self):
        self.root.withdraw()
        n_window = tk.Toplevel()
        n_window.title("Calcular N")
        self.Mitad_pantalla(400,400,n_window)
        n_window.configure(bg='LightSkyBlue')

        vcmd_numeric = (self.root.register(self.validate_numeric), '%P')
        vcmd_rate = (self.root.register(self.validate_rate), '%P')

        # Campo para Valor Presente
        ttk.Label(n_window, text="Valor Presente:", background='LightSkyBlue').pack(pady=5)
        valor_presente_entry = ttk.Entry(n_window, width=30,validate='key', validatecommand=vcmd_numeric)
        valor_presente_entry.pack(pady=5)

        # Campo para Valor Futuro
        ttk.Label(n_window, text="Valor Futuro:", background='LightSkyBlue').pack(pady=5)
        valor_futuro_entry = ttk.Entry(n_window, width=30,validate='key', validatecommand=vcmd_numeric)
        valor_futuro_entry.pack(pady=5)

        # Campo para Tasa
        ttk.Label(n_window, text="Tasa:", background='LightSkyBlue').pack(pady=5)
        tasa_frame = ttk.Frame(n_window)
        tasa_frame.pack(pady=5)
        tasa_entry = ttk.Entry(tasa_frame, width=15,validate='key', validatecommand=vcmd_rate)
        tasa_entry.grid(row=0, column=0, padx=5)

        tipo_tasa = ttk.Combobox(tasa_frame, values=["Efectiva", "Nominal"], state="readonly", width=10)
        tipo_tasa.grid(row=0, column=1, padx=5)

        periodo_tasa = ttk.Combobox(tasa_frame, values=["M", "B", "T", "C", "S", "A"], state="readonly", width=5)
        periodo_tasa.grid(row=0, column=2, padx=5)

        # Campo para Output
        output_label = ttk.Label(n_window, text="", background='LightSkyBlue')
        output_label.pack(pady=10)

        # Botón Calcular
        calculate_button = ttk.Button(n_window, text="Calcular", 
                                    command=lambda: self.Calcular_NICompuesto(valor_presente_entry, valor_futuro_entry, tasa_entry, tipo_tasa, periodo_tasa, output_label))
        calculate_button.pack(pady=10)

        # Botón Volver
        back_button = ttk.Button(n_window, text="Volver", command=lambda: self.Volver(n_window))
        back_button.pack(pady=10)

        n_window.protocol("WM_DELETE_WINDOW", lambda: self.Volver(n_window))

    def CompuestoTasa(self):
        self.root.withdraw()
        tasa_window = tk.Toplevel()
        tasa_window.title("Calcular Tasa")
        self.Mitad_pantalla(400,400,tasa_window)
        tasa_window.configure(bg='LightSkyBlue')

        vcmd_numeric = (self.root.register(self.validate_numeric), '%P')

        # Campo para Valor Presente
        ttk.Label(tasa_window, text="Valor Presente:", background='LightSkyBlue').pack(pady=5)
        valor_presente_entry = ttk.Entry(tasa_window, width=30,validate='key', validatecommand=vcmd_numeric)
        valor_presente_entry.pack(pady=5)

        # Campo para Valor Futuro
        ttk.Label(tasa_window, text="Valor Futuro:", background='LightSkyBlue').pack(pady=5)
        valor_futuro_entry = ttk.Entry(tasa_window, width=30,validate='key', validatecommand=vcmd_numeric)
        valor_futuro_entry.pack(pady=5)

        # Campo para N
        ttk.Label(tasa_window, text="N:", background='LightSkyBlue').pack(pady=5)
        n_frame = ttk.Frame(tasa_window)
        n_frame.pack(pady=5)
        n_entry = ttk.Entry(n_frame, width=15,validate='key', validatecommand=vcmd_numeric)
        n_entry.grid(row=0, column=0, padx=5)

        periodo_n = ttk.Combobox(n_frame, values=["M", "B", "T", "C", "S", "A"], state="readonly", width=5)
        periodo_n.grid(row=0, column=1, padx=5)

        # Campo para Output
        output_label = ttk.Label(tasa_window, text="", background='LightSkyBlue')
        output_label.pack(pady=10)

        # Botón Calcular
        calculate_button = ttk.Button(tasa_window, text="Calcular", 
                                    command=lambda: self.Calcular_CompuestoT(valor_presente_entry, valor_futuro_entry, n_entry, periodo_n, output_label))
        calculate_button.pack(pady=10)

        # Botón Volver
        back_button = ttk.Button(tasa_window, text="Volver", command=lambda: self.Volver(tasa_window))
        back_button.pack(pady=10)

        tasa_window.protocol("WM_DELETE_WINDOW", lambda: self.Volver(tasa_window))

    def ICompuesto(self):
        self.root.withdraw()
        VCompuesto = tk.Toplevel()
        VCompuesto.title("Interés Compuesto")
        self.Mitad_pantalla(400,400,VCompuesto)
        VCompuesto.configure(bg='LightSkyBlue')

        # Create validation commands
        vcmd_numeric = (self.root.register(self.validate_numeric), '%P')
        vcmd_rate = (self.root.register(self.validate_rate), '%P')

        # Combobox for selection
        ttk.Label(VCompuesto, text="Seleccione el cálculo:", background='LightSkyBlue').pack(pady=5)
        calc_type = ttk.Combobox(VCompuesto, values=["ValorPresente", "ValorFuturo"], state="readonly")
        calc_type.pack(pady=5)

        # Create inputs for valor, tasa, and n
        valor_label = ttk.Label(VCompuesto, text="Valor:", background='LightSkyBlue')
        valor_label.pack(pady=5)
        valor_entry = ttk.Entry(VCompuesto, width=30, validate='key', validatecommand=vcmd_numeric)
        valor_entry.pack(pady=5)

        # Tasa section
        ttk.Label(VCompuesto, text="Tasa:", background='LightSkyBlue').pack(pady=5)
        tasa_frame = ttk.Frame(VCompuesto)
        tasa_frame.pack(pady=5)
        tasa_entry = ttk.Entry(tasa_frame, width=15, validate='key', validatecommand=vcmd_rate)
        tasa_entry.grid(row=0, column=0, padx=5)

        tipo_tasa = ttk.Combobox(tasa_frame, values=["Efectiva", "Nominal"], state="readonly", width=10)
        tipo_tasa.grid(row=0, column=1, padx=5)

        periodo_tasa = ttk.Combobox(tasa_frame, values=["M", "B", "T", "C", "S", "A"], state="readonly", width=5)
        periodo_tasa.grid(row=0, column=2, padx=5)

        # Periodos section
        ttk.Label(VCompuesto, text="N:", background='LightSkyBlue').pack(pady=5)
        periodos_frame = ttk.Frame(VCompuesto)
        periodos_frame.pack(pady=5)
        n_entry = ttk.Entry(periodos_frame, width=15, validate='key', validatecommand=vcmd_numeric)
        n_entry.grid(row=0, column=0, padx=5)

        periodo_n = ttk.Combobox(periodos_frame, values=["M", "B", "T", "C", "S", "A"], state="readonly", width=5)
        periodo_n.grid(row=0, column=1, padx=5)

        # Create a space for the output
        output_label = ttk.Label(VCompuesto, text="", background='LightSkyBlue')
        output_label.pack(pady=10)

        CalcularBtn = ttk.Button(VCompuesto, width=15, style='TButton', text="Calcular",command=lambda: self.Calcular_ICompuesto(valor_entry, tasa_entry, tipo_tasa, periodo_tasa, n_entry, periodo_n, calc_type,output_label))

        CalcularBtn.pack(pady=10)        

        VolverBtn = ttk.Button(VCompuesto, width=15, style='TButton', text="Volver",command=lambda: self.Volver(VCompuesto))
        VolverBtn.pack(pady=10)

        VCompuesto.protocol("WM_DELETE_WINDOW", lambda: self.Volver(VCompuesto))

        # Update entries based on selection
        calc_type.bind("<<ComboboxSelected>>", lambda event: self.update_entries(event, calc_type, valor_label))        

    def EQtasas(self):
        self.root.withdraw()
        eq_window = tk.Toplevel()
        eq_window.title("Equivalencia de Tasas")
        self.Mitad_pantalla(600,350,eq_window)
        eq_window.configure(bg='LightSkyBlue')

        # Create validation command
        vcmd_rate = (self.root.register(self.validate_rate), '%P')

        # Campo para Tasa
        ttk.Label(eq_window, text="Tasa:", background='LightSkyBlue').pack(pady=30)
        tasa_frame = ttk.Frame(eq_window)
        tasa_frame.pack(pady=5)
        tasa_entry = ttk.Entry(tasa_frame, width=15, validate='key', validatecommand=vcmd_rate)
        tasa_entry.grid(row=0, column=0, padx=5)

        tipo_tasa = ttk.Combobox(tasa_frame, values=["Efectiva", "Nominal","Efectiva Anticipada","Nominal Anticipada"], state="readonly", width=20)
        tipo_tasa.grid(row=0, column=1, padx=5)

        periodo_tasa = ttk.Combobox(tasa_frame, values=["M", "B", "T", "C", "S", "A"], state="readonly", width=5)
        periodo_tasa.grid(row=0, column=2, padx=5)

        # Label para convertir a
        ttk.Label(eq_window, text="Convertir a:", background='LightSkyBlue').pack(pady=5)

        # Combobox para objetivo_tasa y periodo_tasaO
        objetivo_frame = ttk.Frame(eq_window)
        objetivo_frame.pack(pady=5)
        objetivo_tasa = ttk.Combobox(objetivo_frame, values=["Efectiva", "Nominal","Efectiva Anticipada","Nominal Anticipada"], state="readonly", width=20)
        objetivo_tasa.grid(row=0, column=0, padx=5)

        periodo_tasaO = ttk.Combobox(objetivo_frame, values=["M", "B", "T", "C", "S", "A"], state="readonly", width=5)
        periodo_tasaO.grid(row=0, column=1, padx=5)

        # Campo para Output
        output_label = ttk.Label(eq_window, text="", background='LightSkyBlue')
        output_label.pack(pady=10)

        # Botón Calcular
        calculate_button = ttk.Button(eq_window,width=10, text="Calcular", command=lambda: self.CalcularEQ(tasa_entry, tipo_tasa, periodo_tasa, objetivo_tasa, periodo_tasaO, output_label))
        calculate_button.pack(pady=10)

        # Botón Volver
        VolverBtn = ttk.Button(eq_window, width=10, style='TButton', text="Volver", command=lambda: self.Volver(eq_window))
        VolverBtn.pack(pady=10)

        eq_window.protocol("WM_DELETE_WINDOW", lambda: self.Volver(eq_window))

    def AnualidadMain(self):
        self.root.withdraw()
        window_Amain = tk.Toplevel()
        window_Amain.title("Anualidades")
        self.Mitad_pantalla(400,350,window_Amain)
        window_Amain.configure(bg='LightSkyBlue')

        window_Amain.configure(bg='LightSkyBlue')
        # Combobox para selección de opción
        ttk.Label(window_Amain, text="Seleccione el cálculo:", background='LightSkyBlue').pack(pady=25)
        options = ["Anualidades", "Cuota", "Anualidad Diferida"]
        option_combobox = ttk.Combobox(window_Amain, values=options, state="readonly",width=18)
        option_combobox.pack(pady=5)

        # Botón Continuar
        continue_button = ttk.Button(window_Amain, text="Continuar",width=10, 
                                    command=lambda: self.ValidarOpcionComp(option_combobox.get(), window_Amain))
        continue_button.pack(pady=10)

        VolverBtn = ttk.Button(window_Amain, width=10, style='TButton', text="Volver",command=lambda: self.Volver(window_Amain))
        VolverBtn.pack(pady=10)
        window_Amain.protocol("WM_DELETE_WINDOW", lambda: self.Volver(window_Amain))           

    def Anualidades(self):
        self.root.withdraw()
        anualidades_window = tk.Toplevel()
        anualidades_window.title("Cálculo de Anualidades")
        self.Mitad_pantalla(400, 500, anualidades_window)
        anualidades_window.configure(bg='LightSkyBlue')

        # Create validation commands
        vcmd_rate = (self.root.register(self.validate_rate), '%P')
        vcmd_numeric = (self.root.register(self.validate_numeric), '%P')

        # Combobox para tipo de cálculo
        ttk.Label(anualidades_window, text="Tipo de Cálculo:", background='LightSkyBlue').pack(pady=5)
        calculo_combobox = ttk.Combobox(anualidades_window, values=["Anualidad Vencida S", "Anualidad Vencida a", "Anualidad Anticipada S", "Anualidad Anticipada a"], state="readonly")
        calculo_combobox.pack(pady=5)
        
        # Label que cambia según el cálculo elegido
        descripcion_label = ttk.Label(anualidades_window, text="", background='LightSkyBlue')
        descripcion_label.pack(pady=5)

        def update_label(event):
            seleccion = calculo_combobox.get()
            descripcion_label.config(text=f"Ha seleccionado: {seleccion}")

        calculo_combobox.bind("<<ComboboxSelected>>", update_label)

        # Entry para cuota
        ttk.Label(anualidades_window, text="Cuota:", background='LightSkyBlue').pack(pady=5)
        cuota_entry = ttk.Entry(anualidades_window, width=18, validate='key', validatecommand=vcmd_numeric)
        cuota_entry.pack(pady=5)

        # Tasa section
        ttk.Label(anualidades_window, text="Tasa:", background='LightSkyBlue').pack(pady=5)
        tasa_frame = ttk.Frame(anualidades_window)
        tasa_frame.pack(pady=5)
        tasa_entry = ttk.Entry(tasa_frame, width=15, validate='key', validatecommand=vcmd_rate)
        tasa_entry.grid(row=0, column=0, padx=5)

        tipo_tasa = ttk.Combobox(tasa_frame, values=["Efectiva", "Nominal"], state="readonly", width=10)
        tipo_tasa.grid(row=0, column=1, padx=5)

        periodo_tasa = ttk.Combobox(tasa_frame, values=["M", "B", "T", "C", "S", "A"], state="readonly", width=5)
        periodo_tasa.grid(row=0, column=2, padx=5)

        # Periodos section
        ttk.Label(anualidades_window, text="N:", background='LightSkyBlue').pack(pady=5)
        periodos_frame = ttk.Frame(anualidades_window)
        periodos_frame.pack(pady=5)
        n_entry = ttk.Entry(periodos_frame, width=15, validate='key', validatecommand=vcmd_numeric)
        n_entry.grid(row=0, column=0, padx=5)

        periodo_n = ttk.Combobox(periodos_frame, values=["M", "B", "T", "C", "S", "A"], state="readonly", width=5)
        periodo_n.grid(row=0, column=1, padx=5)

        # Campo para Output
        output_label = ttk.Label(anualidades_window, text="", background='LightSkyBlue')
        output_label.pack(pady=10)

        # Botón Calcular
        calculate_button = ttk.Button(anualidades_window, text="Calcular",width=10, 
                                    command=lambda: self.Calcular_Anualidades(cuota_entry, tasa_entry, tipo_tasa, periodo_tasa, n_entry, periodo_n, calculo_combobox, output_label))
        calculate_button.pack(pady=10)

        # Botón Volver
        VolverBtn = ttk.Button(anualidades_window, width=10, style='TButton', text="Volver", command=lambda: self.Volver(anualidades_window))
        VolverBtn.pack(pady=10)

        anualidades_window.protocol("WM_DELETE_WINDOW", lambda: self.Volver(anualidades_window))                   

    def AnualidadesCuota(self):
        self.root.withdraw()
        anualidadescuota_window = tk.Toplevel()
        anualidadescuota_window.title("Cálculo de Anualidades")
        self.Mitad_pantalla(400, 500, anualidadescuota_window)
        anualidadescuota_window.configure(bg='LightSkyBlue')

        # Create validation commands
        vcmd_rate = (self.root.register(self.validate_rate), '%P')
        vcmd_numeric = (self.root.register(self.validate_numeric), '%P')

        # Combobox para tipo de cálculo
        ttk.Label(anualidadescuota_window, text="Tipo de Cálculo:", background='LightSkyBlue').pack(pady=20)
        calculo_combobox = ttk.Combobox(anualidadescuota_window, values=["Anualidad Vencida S", "Anualidad Vencida a", "Anualidad Anticipada S", "Anualidad Anticipada a"], state="readonly")
        calculo_combobox.pack(pady=5)

        # Entry para valor
        valor_label = ttk.Label(anualidadescuota_window, text="", background='LightSkyBlue')
        valor_label.pack(pady=5)
        valor_entry = ttk.Entry(anualidadescuota_window, width=30, validate='key', validatecommand=vcmd_numeric)
        valor_entry.pack(pady=5)

        def update_label(event):
            seleccion = calculo_combobox.get()
            if seleccion == "Anualidad Vencida S":
                valor_label.config(text="Valor Futuro Vencido :")
            elif seleccion == "Anualidad Vencida a":
                valor_label.config(text="Valor Presente Vencido :")
            elif seleccion == "Anualidad Anticipada S":
                valor_label.config(text="Valor Futuro Anticipado :")
            elif seleccion == "Anualidad Anticipada a":
                valor_label.config(text="Valor Presente Anticipado :")

        calculo_combobox.bind("<<ComboboxSelected>>", update_label)


        # Tasa section
        ttk.Label(anualidadescuota_window, text="Tasa:", background='LightSkyBlue').pack(pady=5)
        tasa_frame = ttk.Frame(anualidadescuota_window)
        tasa_frame.pack(pady=5)
        tasa_entry = ttk.Entry(tasa_frame, width=15, validate='key', validatecommand=vcmd_rate)
        tasa_entry.grid(row=0, column=0, padx=5)

        tipo_tasa = ttk.Combobox(tasa_frame, values=["Efectiva", "Nominal"], state="readonly", width=10)
        tipo_tasa.grid(row=0, column=1, padx=5)

        periodo_tasa = ttk.Combobox(tasa_frame, values=["M", "B", "T", "C", "S", "A"], state="readonly", width=5)
        periodo_tasa.grid(row=0, column=2, padx=5)

        # Periodos section
        ttk.Label(anualidadescuota_window, text="N:", background='LightSkyBlue').pack(pady=5)
        periodos_frame = ttk.Frame(anualidadescuota_window)
        periodos_frame.pack(pady=5)
        n_entry = ttk.Entry(periodos_frame, width=15, validate='key', validatecommand=vcmd_numeric)
        n_entry.grid(row=0, column=0, padx=5)

        periodo_n = ttk.Combobox(periodos_frame, values=["M", "B", "T", "C", "S", "A"], state="readonly", width=5)
        periodo_n.grid(row=0, column=1, padx=5)

        # Campo para Output
        output_label = ttk.Label(anualidadescuota_window, text="", background='LightSkyBlue')
        output_label.pack(pady=10)

        # Botón Calcular
        calculate_button = ttk.Button(anualidadescuota_window, text="Calcular",width=10,
                                    command=lambda: self.Calcular_CuotaAnualidades(valor_entry, tasa_entry, tipo_tasa, periodo_tasa, n_entry, periodo_n, calculo_combobox, output_label))
        calculate_button.pack(pady=10)

        # Botón Volver
        VolverBtn = ttk.Button(anualidadescuota_window, width=10, style='TButton', text="Volver", command=lambda: self.Volver(anualidadescuota_window))
        VolverBtn.pack(pady=10)

        anualidadescuota_window.protocol("WM_DELETE_WINDOW", lambda: self.Volver(anualidadescuota_window))                   

    def Anualidad_Diferida(self):
        self.root.withdraw()
        anualidadesdif_window = tk.Toplevel()
        anualidadesdif_window.title("Cálculos de Anualidades Diferidas")
        self.Mitad_pantalla(400, 500, anualidadesdif_window)
        anualidadesdif_window.configure(bg='LightSkyBlue')

        # Create validation commands
        vcmd_rate = (self.root.register(self.validate_rate), '%P')
        vcmd_numeric = (self.root.register(self.validate_numeric), '%P')

        # Combobox para tipo de cálculo
        ttk.Label(anualidadesdif_window, text="Tipo de Cálculo:", background='LightSkyBlue').pack(pady=25)
        calculo_combobox = ttk.Combobox(anualidadesdif_window, values=["Anualidad Diferida", "Cuota"], state="readonly")
        calculo_combobox.pack(pady=5)
        
        # Entry para valor
        valor_label = ttk.Label(anualidadesdif_window, text="", background='LightSkyBlue')
        valor_label.pack(pady=5)
        valor_entry = ttk.Entry(anualidadesdif_window, width=30, validate='key', validatecommand=vcmd_numeric)
        valor_entry.pack(pady=5)

        def update_label(event):
            seleccion = calculo_combobox.get()
            if seleccion == "Anualidad Diferida":
                valor_label.config(text="Cuota")
            elif seleccion == "Cuota":
                valor_label.config(text="Valor Presente Anticipado")

        calculo_combobox.bind("<<ComboboxSelected>>", update_label)
        
        # Tasa section
        ttk.Label(anualidadesdif_window, text="Tasa:", background='LightSkyBlue').pack(pady=5)
        tasa_frame = ttk.Frame(anualidadesdif_window)
        tasa_frame.pack(pady=5)
        tasa_entry = ttk.Entry(tasa_frame, width=15, validate='key', validatecommand=vcmd_rate)
        tasa_entry.grid(row=0, column=0, padx=5)

        tipo_tasa = ttk.Combobox(tasa_frame, values=["Efectiva", "Nominal"], state="readonly", width=10)
        tipo_tasa.grid(row=0, column=1, padx=5)

        periodo_tasa = ttk.Combobox(tasa_frame, values=["M", "B", "T", "C", "S", "A"], state="readonly", width=5)
        periodo_tasa.grid(row=0, column=2, padx=5)

        # Periodos section
        ttk.Label(anualidadesdif_window, text="N:", background='LightSkyBlue').pack(pady=5)
        periodos_frame = ttk.Frame(anualidadesdif_window)
        periodos_frame.pack(pady=5)
        n_entry = ttk.Entry(periodos_frame, width=15, validate='key', validatecommand=vcmd_numeric)
        n_entry.grid(row=0, column=0, padx=5)

        periodo_n = ttk.Combobox(periodos_frame, values=["M", "B", "T", "C", "S", "A"], state="readonly", width=5)
        periodo_n.grid(row=0, column=1, padx=5)

        # Entry para m y combobox pm
        ttk.Label(anualidadesdif_window, text="M:", background='LightSkyBlue').pack(pady=5)
        m_frame = ttk.Frame(anualidadesdif_window)
        m_frame.pack(pady=5)
        m_entry = ttk.Entry(m_frame, width=15, validate='key', validatecommand=vcmd_numeric)
        m_entry.grid(row=0, column=0, padx=5)

        pm = ttk.Combobox(m_frame, values=["M", "B", "T", "C", "S", "A"], state="readonly", width=5)
        pm.grid(row=0, column=1, padx=5)

        # Campo para Output
        output_label = ttk.Label(anualidadesdif_window, text="", background='LightSkyBlue')
        output_label.pack(pady=10)

        # Botón Calcular
        calculate_button = ttk.Button(anualidadesdif_window, text="Calcular", width=10, 
                                    command=lambda: self.Calcular_AnualidadesDif(valor_entry, tasa_entry, tipo_tasa, periodo_tasa, n_entry, periodo_n, m_entry, pm, calculo_combobox, output_label))
        calculate_button.pack(pady=10)

        # Botón Volver
        VolverBtn = ttk.Button(anualidadesdif_window, width=10, style='TButton', text="Volver", command=lambda: self.Volver(anualidadesdif_window))
        VolverBtn.pack(pady=10)

        anualidadesdif_window.protocol("WM_DELETE_WINDOW", lambda: self.Volver(anualidadesdif_window))

    def Tables(self):
        self.root.withdraw()
        tables_window = tk.Toplevel()
        tables_window.title("Tables")
        self.Mitad_pantalla(400, 350, tables_window)
        tables_window.configure(bg='LightSkyBlue')

        # Create validation commands
        vcmd_abono = (self.root.register(self.validate_abono), '%P')

        # Label para tipo de tabla
        ttk.Label(tables_window, text="Tipo de Tabla:", background='LightSkyBlue').pack(pady=5)
        tipo_tabla_combobox = ttk.Combobox(tables_window, values=["Amortización", "Capitalización"], state="readonly")
        tipo_tabla_combobox.pack(pady=5)

        # Entry para abono
        ttk.Label(tables_window, text="Abono:", background='LightSkyBlue').pack(pady=5)
        abono_frame = ttk.Frame(tables_window)
        abono_frame.pack(pady=5)
        abono_entry = ttk.Entry(abono_frame, width=15, validate='key', validatecommand=vcmd_abono)
        abono_entry.grid(row=0, column=0, padx=5)

        # Combobox para periodo de abono, valores dependen de self.n
        periodo_combobox = ttk.Combobox(abono_frame, values=list(range(self.n + 1)), state="readonly", width=10)
        periodo_combobox.grid(row=0, column=1, padx=5)

        # Botón Generar Tabla
        generar_button = ttk.Button(tables_window, text="Generar Tabla", width=15, command=lambda: self.GenerarTabla(tipo_tabla_combobox, abono_entry, periodo_combobox))
        generar_button.pack(pady=10)

        # Botón Volver
        volver_button = ttk.Button(tables_window, text="Volver", width=15, command=lambda: self.Volver(tables_window))
        volver_button.pack(pady=10)

        tables_window.protocol("WM_DELETE_WINDOW", lambda: self.Volver(tables_window))      

    def validate_numeric(self,P):
        try:
            if P == "":
                return True  # Permitir que el campo esté vacío temporalmente
            value = float(P)
            return value >= 0  # Permitir solo números positivos
        except ValueError:
            return False
        
    def validate_abono(self, P):
        if P == "":
            return True  # Permitir que el campo esté vacío temporalmente
        try:
            value = float(P)
            return 0 <= value <= self.a  # Validar que el valor esté entre 0 y self.a
        except ValueError:
            return False        
          
    def validate_rate(self, P):
        try:
            if P == "":
                return True
            value = float(P)
            if 0.0 <= value <= 1.0:
                return True
            else:
                return False
        except ValueError:
            return False
        
    def on_closing(self):
        self.root.quit()

    def update_entries(self, event, calc_type, valor_label):
        calc_type_value = calc_type.get()
        if calc_type_value == "ValorPresente":
            valor_label.config(text="Valor Futuro:")
        elif calc_type_value == "ValorFuturo":
            valor_label.config(text="Valor Presente:")

    def Volver(self,ventana):
        ventana.withdraw()
        ventana.destroy()
        self.root.deiconify()

    def Calcular_ICompuesto(self,valor,tasa,ttasa,ptasa,n,pn,tipo,output):

        if not valor.get() or not tasa.get() or not ttasa.get() or not ptasa.get() or not n.get() or not pn.get() or not tipo.get():
            output.configure(text="Todos los valores deben ser ingresados")
        else:
            if ttasa.get() == "Nominal" :
                if ptasa.get() == "M":
                    tasa = convertirJI(0,float(tasa.get()),2,12)
                elif ptasa.get() == "B":
                    tasa = convertirJI(0,float(tasa.get()),2,6)             
                elif ptasa.get() == "T":
                    tasa = convertirJI(0,float(tasa.get()),2,4)          
                elif ptasa.get() == "C":
                    tasa = convertirJI(0,float(tasa.get()),2,3)              
                elif ptasa.get() == "S":
                    tasa = convertirJI(0,float(tasa.get()),2,2)            
                elif ptasa.get() == "A":
                    tasa = convertirJI(0,float(tasa.get()),2,1)

                n = Cambiar_n(int(n.get()),ptasa,pn)

                if tipo.get() == "ValorPresente":
                    resultado=(f"El valor presente es = {interesCompuestoP(float(valor.get()),tasa,n)} en periodos {ptasa.get()}")
                    output.config(text=resultado)                
                else:
                    resultado=(f"El valor futuro es = {interesCompuesto(float(valor.get()),tasa,n)} en periodos {ptasa.get()}")
                    output.config(text=resultado)                 
            else:
                n = Cambiar_n(int(n.get()),ptasa,pn)

                if tipo.get() == "ValorPresente":
                    resultado=(f"El valor presente es = {interesCompuestoP(float(valor.get()),float(tasa.get()),n)} en periodos {ptasa.get()}")
                    output.config(text=resultado)                
                else:
                    resultado=(f"El valor futuro es = {interesCompuesto(float(valor.get()),float(tasa.get()),n)} en periodos {ptasa.get()}")
                    output.config(text=resultado)            

    def ValidarOpcionComp(self,option,mainwindow):
        if option == "valor":
            mainwindow.withdraw()
            self.ICompuesto()
        elif option =="tiempo":
            mainwindow.withdraw()
            self.CompuestoN()
        elif option == "tasa":
            mainwindow.withdraw()
            self.CompuestoTasa()
        elif option == "Anualidades":
            mainwindow.withdraw()
            self.Anualidades()
        elif option == "Cuota":
            mainwindow.withdraw()
            self.AnualidadesCuota()
        elif option == "Anualidad Diferida":
            mainwindow.withdraw()
            self.Anualidad_Diferida()

    def Mitad_pantalla(self,width,heigth,main_window):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        top_level_width = width
        top_level_height = heigth 
        x = (screen_width - top_level_width) // 2
        y = (screen_height - top_level_height) // 2
        main_window.geometry(f"{top_level_width}x{top_level_height}+{x}+{y}")

    def Calcular_NICompuesto(self,valor_presente, valor_futuro, tasa, tipo_tasa, ptasa, output):

        if not valor_presente.get() or not valor_futuro.get() or not tasa.get() or not tipo_tasa.get() or not ptasa.get():
            output.config(text="Todos los valores deben ser ingresados")
        else:
            if tipo_tasa.get() =='Nominal':
                if ptasa.get() == "M":
                    tasa = convertirJI(0,float(tasa.get()),2,12)
                elif ptasa.get() == "B":
                    tasa = convertirJI(0,float(tasa.get()),2,6)             
                elif ptasa.get() == "T":
                    tasa = convertirJI(0,float(tasa.get()),2,4)          
                elif ptasa.get() == "C":
                    tasa = convertirJI(0,float(tasa.get()),2,3)              
                elif ptasa.get() == "S":
                    tasa = convertirJI(0,float(tasa.get()),2,2)            
                elif ptasa.get() == "A":
                    tasa = convertirJI(0,float(tasa.get()),2,1)

                respuesta = (f'El valor de n es = {interesCompuestoN(float(valor_futuro.get()),float(valor_presente.get()),tasa)} en periodos de {ptasa.get()}')
                output.config(text=respuesta)
            else:
                respuesta = (f'El valor de n es = {interesCompuestoN(float(valor_futuro.get()),float(valor_presente.get()),float(tasa.get()))} en periodos de {ptasa.get()}')
                output.config(text=respuesta)
    
    def Calcular_CompuestoT(self,valor_presente,valor_futuro,n,pn, output):

        if not valor_presente.get() or not valor_futuro.get() or not n.get() or not pn.get():
            output.config(text="Todos los campos deben ser ingresados")
        else:
            respuesta = (f'El valor de la tasa es de {interesCompuestoI(float(valor_futuro.get()),float(valor_presente.get()),int(n.get()))} en periodos {pn.get()}')
            output.config(text=respuesta)        
            
    def CalcularEQ(self,tasa_entry, tipo_tasa, periodo_tasa, objetivo_tasa, periodo_tasaO, output_label):
        conversiones = {
            'M': 12,  # Mensual
            'B': 6,   # Bimestral
            'T': 4,   # Trimestral
            'C': 3,   # Cuatrimestral
            'S': 2,   # Semestral
            'A': 1    # Anual
        }
        if not tasa_entry.get() or not tipo_tasa.get() or not periodo_tasa.get() or not objetivo_tasa.get() or not periodo_tasaO.get():
            output_label.config(text="Todas las entradas deben indicarse")
        else:
            if tipo_tasa.get() == 'Nominal Anticipada':
                tasa = convertirJI(0,float(tasa_entry.get()),2,conversiones[periodo_tasa.get()])
                tasa = convertirIA(0,tasa,2)
                tasaO = equivalenciaTasas(tasa,conversiones[periodo_tasa.get()],conversiones[periodo_tasaO.get()])
                if objetivo_tasa.get() == 'Nominal Anticipada':
                    tasaO = convertirIA(tasaO,0,1)
                    tasaO = convertirJI(tasaO,0,1,conversiones[periodo_tasaO.get()])
                elif objetivo_tasa.get() == 'Nominal':
                    tasaO = convertirJI(tasaO,0,1,conversiones[periodo_tasaO.get()])
                elif objetivo_tasa.get() == 'Efectiva Anticipada':
                    tasaO = convertirIA(tasaO,0,1)                    
                output_label.config(text=f"La equivalencia de la tasa {tasa_entry.get()} {tipo_tasa.get()} {periodo_tasa.get()} es {tasaO} {objetivo_tasa.get()} {periodo_tasaO.get()}")
            elif tipo_tasa.get() == 'Nominal':                
                tasa = convertirJI(0,float(tasa_entry.get()),2,conversiones[periodo_tasa.get()])
                tasaO = equivalenciaTasas(tasa,conversiones[periodo_tasa.get()],conversiones[periodo_tasaO.get()])
                if objetivo_tasa.get() == "Nominal":
                    tasaO = convertirJI(tasaO,0,1,conversiones[periodo_tasaO.get()])
                elif objetivo_tasa.get() == 'Efectiva Anticipada':
                    tasaO = convertirIA(tasaO,0,1)
                elif objetivo_tasa.get()== 'Nominal Anticipada':
                    tasaO = convertirIA(tasaO,0,1)
                    tasaO = convertirJI(tasaO,0,1,conversiones[periodo_tasaO.get()])
                output_label.config(text=f"La equivalencia de la tasa {tasa_entry.get()} {tipo_tasa.get()} {periodo_tasa.get()} es {tasaO} {objetivo_tasa.get()} {periodo_tasaO.get()}")
            elif tipo_tasa.get() == 'Efectiva Anticipada':
                tasa = convertirIA(0,float(tasa_entry.get()),2)
                tasaO = equivalenciaTasas(tasa,conversiones[periodo_tasa.get()],conversiones[periodo_tasaO.get()])
                if objetivo_tasa.get() == "Nominal":
                    tasaO = convertirJI(tasaO,0,1,conversiones[periodo_tasaO.get()])
                elif objetivo_tasa.get() == 'Efectiva Anticipada':
                    tasaO = convertirIA(tasaO,0,1)
                elif objetivo_tasa.get() == 'Nominal Anticipada':
                    tasaO = convertirIA(tasaO,0,1)
                    tasaO = convertirJI(tasaO,0,1,conversiones[periodo_tasaO.get()])
                output_label.config(text=f"La equivalencia de la tasa {tasa_entry.get()} {tipo_tasa.get()} {periodo_tasa.get()} es {tasaO} {objetivo_tasa.get()} {periodo_tasaO.get()}")
            else:
                tasaO = equivalenciaTasas(float(tasa_entry.get()),conversiones[periodo_tasa.get()],conversiones[periodo_tasaO.get()])
                if objetivo_tasa.get() == "Nominal":
                    tasaO = convertirJI(tasaO,0,1,conversiones[periodo_tasaO.get()])
                elif objetivo_tasa.get() == 'Efectiva Anticipada':
                    tasaO = convertirIA(tasaO,0,1)
                elif objetivo_tasa.get() == 'Nominal Anticipada':
                    tasaO = convertirIA(tasaO,0,1)
                    tasaO = convertirJI(tasaO,0,1,conversiones[periodo_tasaO.get()])
                output_label.config(text=f"La equivalencia de la tasa {tasa_entry.get()} {tipo_tasa.get()} {periodo_tasa.get()} es {tasaO} {objetivo_tasa.get()} {periodo_tasaO.get()}")

    def Calcular_Anualidades(self,cuota, tasa, ttasa, ptasa, n, pn, option, output):
        conversiones = {
            'M': 12,  # Mensual
            'B': 6,   # Bimestral
            'T': 4,   # Trimestral
            'C': 3,   # Cuatrimestral
            'S': 2,   # Semestral
            'A': 1    # Anual
        }
        if not cuota.get() or not tasa.get() or not ttasa.get() or not ptasa.get() or not n.get() or not pn.get() or not option.get():
            output.config(text="Todos los valores deben ser ingresados")
        else:
            if option.get() == "Anualidad Vencida S":
                if ttasa.get() == "Nominal" :
                    tasa = convertirJI(0,float(tasa.get()),2,conversiones[ptasa.get()])
                else:
                    tasa = float(tasa.get())
                n = Cambiar_n(int(n.get()),ptasa,pn)
                valorFuturo = AnualidadOS(float(cuota.get()),tasa,n)
                output.config(text=f"El valor futuro es {valorFuturo}")
                
                self.S =valorFuturo
                self.n = int(n)
                self.i = tasa
                self.A = float(cuota.get())
                self.banderaS = True
                self.banderaa = False
                
            elif option.get() == "Anualidad Vencida a":
                if ttasa.get() == "Nominal" :
                    tasa = convertirJI(0,float(tasa.get()),2,conversiones[ptasa.get()])
                else:
                    tasa = float(tasa.get())
                n = Cambiar_n(int(n.get()),ptasa,pn)
                valorPresente = AnualidadOa(float(cuota.get()),tasa,n)
                output.config(text=f"El valor presente es {valorPresente}")
                
                self.A=float(cuota.get())
                self.a=valorPresente
                self.n = int(n)
                self.i = tasa
                self.banderaa = True
                self.banderaS = False

            elif option.get() == "Anualidad Anticipada S":
                if ttasa.get() == "Nominal" :
                    tasa = convertirJI(0,float(tasa.get()),2,conversiones[ptasa.get()])
                else:
                    tasa = float(tasa.get())
                n = Cambiar_n(int(n.get()),ptasa,pn)
                valorFuturo = AnualidadAS(float(cuota.get()),tasa,n)
                output.config(text=f"El valor futuro es {valorFuturo}")

                self.S =valorFuturo
                self.n = int(n)
                self.i = tasa
                self.A = float(cuota.get())
                self.banderaS = True
                self.banderaa = False

            elif option.get() == "Anualidad Anticipada a":
                if ttasa.get() == "Nominal" :
                    tasa = convertirJI(0,float(tasa.get()),2,conversiones[ptasa.get()])
                else:
                    tasa = float(tasa.get())
                n = Cambiar_n(int(n.get()),ptasa,pn)
                valorPresente = AnualidadAa(float(cuota.get()),tasa,n)
                output.config(text=f"El valor presente es {valorPresente}")
                
                self.A=float(cuota.get())
                self.a=valorPresente
                self.n = int(n)
                self.i = tasa
                self.banderaa = True
                self.banderaS = False

    def Calcular_CuotaAnualidades(self,valor, tasa, ttasa, ptasa, n, pn, option, output):
        conversiones = {
            'M': 12,  # Mensual
            'B': 6,   # Bimestral
            'T': 4,   # Trimestral
            'C': 3,   # Cuatrimestral
            'S': 2,   # Semestral
            'A': 1    # Anual
        }
        if not valor.get() or not tasa.get() or not ttasa.get() or not ptasa.get() or not n.get() or not pn.get() or not option.get():
            output.config(text="Todos los campos deben ser llenados")
        else:
            if option.get() == "Anualidad Vencida S":
                if ttasa.get() == "Nominal" :
                    tasa = convertirJI(0,float(tasa.get()),2,conversiones[ptasa.get()])
                else:
                    tasa = float(tasa.get())
                n = Cambiar_n(int(n.get()),ptasa,pn)
                cuota = AnualidadOS_A(float(valor.get()),tasa,n)
                output.config(text=f"El valor de la cuota es {cuota}")

                self.S = float(valor.get())
                self.n = int(n)
                self.i = tasa
                self.A = cuota
                self.banderaS = True
                self.banderaa = False                                
            elif option.get() == "Anualidad Vencida a":   
                if ttasa.get() == "Nominal" :
                    tasa = convertirJI(0,float(tasa.get()),2,conversiones[ptasa.get()])
                else:
                    tasa = float(tasa.get())
                n = Cambiar_n(int(n.get()),ptasa,pn)
                cuota = AnualidadOa_A(float(valor.get()),tasa,n)
                output.config(text=f"El valor de la cuota es {cuota}")  

                self.A=cuota
                self.a=float(valor.get())
                self.n = int(n)
                self.i = tasa 
                self.banderaa = True
                self.banderaS = False

            elif option.get() == "Anualidad Anticipada S":   
                if ttasa.get() == "Nominal" :
                    tasa = convertirJI(0,float(tasa.get()),2,conversiones[ptasa.get()])
                else:
                    tasa = float(tasa.get())
                n = Cambiar_n(int(n.get()),ptasa,pn)
                cuota = AnualidadAS_A(float(valor.get()),tasa,n)
                output.config(text=f"El valor de la cuota es {cuota}")

                self.S = float(valor.get())
                self.n = int(n)
                self.i = tasa
                self.A = cuota
                self.banderaS = True
                self.banderaa = False

            elif option.get() == "Anualidad Anticipada a":   
                if ttasa.get() == "Nominal" :
                    tasa = convertirJI(0,float(tasa.get()),2,conversiones[ptasa.get()])
                else:
                    tasa = float(tasa.get())
                n = Cambiar_n(int(n.get()),ptasa,pn)
                cuota = AnualidadAa_A(float(valor.get()),tasa,n)
                output.config(text=f"El valor de la cuota es {cuota}")

                self.A=cuota
                self.a=float(valor.get())
                self.n = int(n)
                self.i = tasa
                self.banderaa = True
                self.banderaS = False

    def Calcular_AnualidadesDif(self,valor, tasa, ttasa, ptasa, n, pn, m, pm, option, output):
        conversiones = {
            'M': 12,  # Mensual
            'B': 6,   # Bimestral
            'T': 4,   # Trimestral
            'C': 3,   # Cuatrimestral
            'S': 2,   # Semestral
            'A': 1    # Anual
        }        
        if not valor.get() or not tasa.get() or not ttasa.get() or not ptasa.get() or not n.get() or not pn.get() or not m.get() or not pm.get() or not option.get():
            output.config(text="Todos los campos deben ser llenados")
        else:
            if option.get() == "Anualidad Diferida":                
                if ttasa.get() == "Nominal" :
                    tasa = convertirJI(0,float(tasa.get()),2,conversiones[ptasa.get()])
                else:
                    tasa = float(tasa.get())
                n = Cambiar_n(int(n.get()),ptasa,pn)
                m = Cambiar_n(int(m.get()),ptasa,pm)
                valorpresente = AnualidadDiferida(float(valor.get()),tasa,n,m-1)
                output.config(text=f"El valor es {valorpresente}")

                self.A=float(valor.get())
                self.a=valorpresente
                self.n = int(n)
                self.i = tasa
                self.banderaa = True
                self.banderaS = False

            elif option.get() == "Cuota":
                if ttasa.get() == "Nominal" :
                    tasa = convertirJI(0,float(tasa.get()),2,conversiones[ptasa.get()])
                else:
                    tasa = float(tasa.get())
                n = Cambiar_n(int(n.get()),ptasa,pn)
                m = Cambiar_n(int(m.get()),ptasa,pm)
                cuota = AnualidadDiferidaA(float(valor.get()),tasa,n,m-1)
                output.config(text=f"La cuota  es de {cuota} en periodos {ptasa.get()}")

                self.A=cuota
                self.a=float(valor.get())
                self.n = int(n)
                self.i = tasa 
                self.banderaa = True
                self.banderaS = False                               

    def GenerarTabla(self,tipo_tabla, abono, pabono):        

        if tipo_tabla.get() == "Amortización":
            if self.banderaa == False and self.banderaS==False:
                messagebox.showerror("Error", "No puedes Crear una tabla de amortización sin valores generados por medio de valores presentes")
            elif self.banderaa==False and self.banderaS==True:
                messagebox.showerror("Error", "No puedes Crear una tabla de amortización desde valores futuros")                
            elif self.banderaa == True and self.banderaS ==False:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Tabla Dinámica"

                # Definir las columnas
                columnas = ["Periodo", "Saldo", "Interés", "Cuota", "Amortización"]
                
                # Escribir las columnas en la primera fila
                for col_num, columna in enumerate(columnas, 1):
                    ws.cell(row=1, column=col_num, value=columna)
                    ws.cell(row=1, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

                auxa = self.a 
                Amort = 0
                ab= False            
                # Escribir las filas basadas en n
                for row_num in range(2, int(self.n) + 3):
                    if row_num > 2:
                        if row_num-2 == int(pabono.get()):
                            print('entre a abono')
                            intereses = auxa*self.i
                            nuevacuota = self.A+float(abono.get())                
                            Amort = nuevacuota-intereses
                            auxa-=Amort
                            ab = True
                        else:
                            ab=False
                            intereses = auxa*self.i
                            Amort=self.A-intereses 
                            auxa= auxa-Amort
                    for col_num in range(1, 6):
                        if row_num == 2:
                            if(row_num == 2 and col_num == 2):                                
                                ws.cell(row=row_num, column=col_num, value=self.a)
                                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center") 
                            else:
                                ws.cell(row=row_num, column=col_num, value=0)
                                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")                                                    
                        else:
                            if col_num == 1:
                                ws.cell(row=row_num, column=col_num, value=f"{row_num-2}")
                                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")                            
                            elif col_num == 2:
                                ws.cell(row=row_num, column=col_num, value=f"{auxa}")
                                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")                        
                            elif col_num == 3:
                                ws.cell(row=row_num, column=col_num, value=f"{intereses}")
                                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")
                            elif col_num == 4:
                                if ab == True:
                                    ws.cell(row=row_num, column=col_num, value=f"{self.A+float(abono.get())}")
                                    ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")
                                else:
                                    ws.cell(row=row_num, column=col_num, value=f"{self.A}")
                                    ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")
                            elif col_num == 5:                            
                                ws.cell(row=row_num, column=col_num, value=f"{Amort}")
                                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

                # Guardar el archivo Excel
                archivo = f"Tabla_{tipo_tabla.get()}.xlsx"
                wb.save(archivo)
                print(f"Archivo {archivo} generado con éxito.")
                
                # Opcional: Abrir el archivo automáticamente después de crearlo
                import os
                os.startfile(archivo)
        else:
            if self.banderaa == False and self.banderaS==False:
                messagebox.showerror("Error", "No puedes Crear una tabla de amortización sin valores generados por medio de valores futuros")
            elif self.banderaa==True and self.banderaS==False:
                messagebox.showerror("Error", "No puedes Crear una tabla de capitalizción desde valores presentes")                
            elif self.banderaa == False and self.banderaS ==True:                        
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Tabla Dinámica"

                # Definir las columnas
                columnas = ["Periodo", "Saldo", "Interés", "Cuota", "Incremento"]
                
                # Escribir las columnas en la primera fila
                for col_num, columna in enumerate(columnas, 1):
                    ws.cell(row=1, column=col_num, value=columna)
                    ws.cell(row=1, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

                # Escribir las filas basadas en n
                incremento = self.A
                auxS=self.A
                intereses = auxS

                for row_num in range(2, int(self.n) + 2):
                    if row_num > 2:
                        intereses = auxS*self.i
                        incremento=self.A+intereses 
                        auxS+= incremento
                    for col_num in range(1, 6):
                        if row_num == 2:
                            if col_num == 3:                                                            
                                ws.cell(row=row_num, column=col_num, value=None)
                                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center") 
                            elif col_num == 1:
                                ws.cell(row=row_num, column=col_num, value=1)
                                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")                                                           
                            elif col_num ==2 or col_num>3:
                                ws.cell(row=row_num, column=col_num, value=self.A)
                                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")                                                    
                        else:
                            if col_num == 1:
                                ws.cell(row=row_num, column=col_num, value=f"{row_num-1}")
                                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")                            
                            elif col_num == 2:
                                ws.cell(row=row_num, column=col_num, value=f"{auxS}")
                                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")                        
                            elif col_num == 3:
                                ws.cell(row=row_num, column=col_num, value=f"{intereses}")
                                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")
                            elif col_num == 4:
                                ws.cell(row=row_num, column=col_num, value=f"{self.A}")
                                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")
                            elif col_num == 5:                            
                                ws.cell(row=row_num, column=col_num, value=f"{incremento}")
                                ws.cell(row=row_num, column=col_num).alignment = Alignment(horizontal="center", vertical="center")

                # Guardar el archivo Excel
                archivo = f"Tabla_{tipo_tabla.get()}.xlsx"
                wb.save(archivo)
                print(f"Archivo {archivo} generado con éxito.")
                
                # Opcional: Abrir el archivo automáticamente después de crearlo
                import os
                os.startfile(archivo)
